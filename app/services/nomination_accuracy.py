"""Nomination accuracy — replicates ARECO Daily Trading Summary Compliance logic.

MQ: Daily MIRF WESM file, range B13:Y36 (24 rows × DEL/REC pairs → 288 DEL MW).
Compliance: MPI CSV — Interval End, Market DOT (RTD), Actual Output; sorted by time;
  rows mapped to 5‑minute slots (00:05 … 23:55) like Trading Report paste C63:D216.

Backfill: ``RTD -Actual -Day Ahead`` workbook (C–E for times, RTD, Actual) plus optional
  MIRF Daily MQ workbook. When MQ is provided, E uses MIRF DEL (same as MPI+MIRF). Otherwise
  column F (Day Ahead MW) is used as MQ.

FPE (per Excel Compliance col I): ABS((G - E) / H_max) with
  E = MQ (MW), G = (F + RTD) / 24, F = lagged RTD (F[0]=0, F[i]=RTD[i-1]),
  H_max = max(E over day).
MAPE: mean(FPE) over intervals with numeric RTD (COUNT(C) in template).
PERC95: linear interpolation on sorted FPE (p=0.95, h=p*(N+1)).
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, time
from typing import Any

from openpyxl import load_workbook

N_INTERVALS = 288
MQ_ROWS = 24
MQ_COL_START = 2  # B
MQ_COL_END = 25  # Y inclusive (12 DEL + 12 REC pairs)
MQ_SOURCE_FIRST_ROW = 13

# Policy (ARECO): MAPE must stay below 18%; PERC95 must stay below 30%.
MAPE_MAX_EXCLUSIVE = 0.18
PERC95_MAX_EXCLUSIVE = 0.30
N_HOURS = 24
INTERVALS_PER_HOUR = 12


def _interval_index_from_end(t: time) -> int | None:
    """Map interval label/end time to 0..287 (00:05 → 0)."""
    m = t.hour * 60 + t.minute + t.second / 60.0
    if m % 5.0 > 1e-6:
        return None
    idx = int(round((m - 5) / 5))
    if 0 <= idx < N_INTERVALS:
        return idx
    return None


def load_mq_del_mw_from_xlsx(
    content: bytes,
    sheet_name: str | None = None,
) -> tuple[list[float], str]:
    """Extract 288 DEL interval MW values (kWh/1000) from pasted-equivalent grid B13:Y36."""
    bio = io.BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        sn = sheet_name
        if not sn:
            for n in names:
                if n.endswith("_DEL") and "VISTASOL" in n:
                    sn = n
                    break
            if not sn:
                sn = names[0] if names else ""
        if sn not in names:
            raise ValueError(f"Sheet not found: {sn!r}. Available: {names}")
        sh = wb[sn]
        out: list[float] = []
        for r in range(MQ_SOURCE_FIRST_ROW, MQ_SOURCE_FIRST_ROW + MQ_ROWS):
            for c in range(MQ_COL_START, MQ_COL_END + 1, 2):
                raw = sh.cell(r, c).value
                if raw is None:
                    out.append(0.0)
                else:
                    try:
                        out.append(float(raw) / 1000.0)
                    except (TypeError, ValueError):
                        out.append(0.0)
        if len(out) != N_INTERVALS:
            raise ValueError(f"Expected {N_INTERVALS} MQ intervals, got {len(out)}")
        return out, sn
    finally:
        wb.close()


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def parse_compliance_csv(content: bytes) -> list[tuple[datetime, float, float]]:
    """Return sorted (interval_end, rtd_mw, actual_mw)."""
    text = content.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.reader(f)
    rows = list(reader)
    if not rows:
        return []
    header = [_normalize_header(x) for x in rows[0]]
    # find columns
    def pick(*candidates: str) -> int:
        for i, h in enumerate(header):
            for cand in candidates:
                if cand in h or h in cand:
                    return i
        return -1

    i_time = pick("interval end", "interval", "end")
    i_rtd = pick("market dot", "dot", "rtd")
    i_act = pick("actual output", "actual")
    if i_time < 0:
        i_time = 0
    if i_rtd < 0 or i_act < 0:
        raise ValueError(
            "Could not find RTD / Actual columns (expected headers like 'Market DOT' and 'Actual Output')."
        )

    parsed: list[tuple[datetime, float, float]] = []
    for row in rows[1:]:
        if len(row) <= max(i_time, i_rtd, i_act):
            continue
        ts_raw = row[i_time].strip()
        if not ts_raw:
            continue
        dt: datetime | None = None
        for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                dt = datetime.strptime(ts_raw, fmt)
                break
            except ValueError:
                continue
        if dt is None:
            continue
        try:
            rtd = float(str(row[i_rtd]).replace(",", "").strip() or 0)
            act = float(str(row[i_act]).replace(",", "").strip() or 0)
        except ValueError:
            continue
        parsed.append((dt, rtd, act))
    parsed.sort(key=lambda x: x[0])
    return parsed


def _dominant_date(rows: list[tuple[datetime, float, float]]) -> date | None:
    if not rows:
        return None
    counts: dict[date, int] = {}
    for dt, _, _ in rows:
        counts[dt.date()] = counts.get(dt.date(), 0) + 1
    return max(counts, key=counts.get)


def fill_rtd_actual(
    compliance: list[tuple[datetime, float, float]],
    day: date,
    t_start: time = time(5, 5),
    t_end: time = time(19, 0),
) -> tuple[list[float], list[float], int]:
    """Return rtd[288], actual[288], count of rows applied in window."""
    rtd = [0.0] * N_INTERVALS
    actual = [0.0] * N_INTERVALS
    n_applied = 0
    for dt, rv, av in compliance:
        if dt.date() != day:
            continue
        t = dt.time()
        if t < t_start or t > t_end:
            continue
        idx = _interval_index_from_end(t)
        if idx is None:
            continue
        rtd[idx] = rv
        actual[idx] = av
        n_applied += 1
    return rtd, actual, n_applied


def _mw_series_stats(xs: list[float]) -> dict[str, float]:
    """Match Summary sheet style: SUM / MIN / AVERAGE / MAX over all 288 intervals."""
    if len(xs) != N_INTERVALS:
        raise ValueError("Expected 288 values")
    s = float(sum(xs))
    return {
        "sum": s,
        "min": float(min(xs)),
        "max": float(max(xs)),
        "mean": s / N_INTERVALS,
    }


def _daily_mwh(mw: list[float]) -> float:
    """Trading Report col G: each interval MWh = MW/12 (5-minute slots)."""
    return float(sum(x / 12.0 for x in mw))


def compute_summary_style_analytics(
    rtd_mw: list[float],
    actual_mw: list[float],
    mq_mw: list[float],
    fpe: list[float | None],
) -> dict[str, Any]:
    """
    Mirrors ``Summary`` / ``Summary_for print`` quantities that only need RTD, Actual, MQ:
    MW totals (like Trading C291-style SUM), min/ave/max rows, daily MWh, and hourly MWh blocks
    (like ``Summary`` row 19 sums of G3:G14, G15:G26, …).
    """
    hourly: list[dict[str, Any]] = []
    for h in range(N_HOURS):
        lo = h * INTERVALS_PER_HOUR
        hi = lo + INTERVALS_PER_HOUR
        hourly.append(
            {
                "hour": h + 1,
                "label": f"Hour {h + 1}",
                "rtd_mwh": float(sum(rtd_mw[lo:hi]) / 12.0),
                "actual_mwh": float(sum(actual_mw[lo:hi]) / 12.0),
                "mq_del_mwh": float(sum(mq_mw[lo:hi]) / 12.0),
            }
        )
    valid_fpe = [x for x in fpe if x is not None]
    return {
        "source": "Matches ARECO Daily Trading Summary: Summary sheet MW/MWh stats; "
        "hourly MWh = SUM(MW/12) per 12×5‑min block like Trading Report column G.",
        "trading_summary_mw": {
            "real_time_dispatch_mw": _mw_series_stats(rtd_mw),
            "actual_dispatch_mw": _mw_series_stats(actual_mw),
            "mq_delivered_mw": _mw_series_stats(mq_mw),
        },
        "trading_summary_mwh_day": {
            "real_time_dispatch_mwh": _daily_mwh(rtd_mw),
            "actual_dispatch_mwh": _daily_mwh(actual_mw),
            "mq_delivered_mwh": _daily_mwh(mq_mw),
        },
        "hourly_mwh": hourly,
        "fpe": {
            "max": float(max(valid_fpe)) if valid_fpe else None,
            "mean": float(sum(valid_fpe) / len(valid_fpe)) if valid_fpe else None,
        },
    }


def evaluate_nomination_policy(mape: float | None, perc95: float | None) -> dict[str, Any]:
    """
    MAPE: non-compliant if >= 18% (must stay below 18%).
    PERC95: non-compliant if >= 30% (must stay below 30%).
    """
    mape_v = mape is not None and mape >= MAPE_MAX_EXCLUSIVE
    p95_v = perc95 is not None and perc95 >= PERC95_MAX_EXCLUSIVE
    mape_ok = mape is not None and mape < MAPE_MAX_EXCLUSIVE
    perc95_ok = perc95 is not None and perc95 < PERC95_MAX_EXCLUSIVE
    day_ok = mape_ok and perc95_ok
    failure_reasons: list[str] = []
    notes: list[str] = []
    if mape is None:
        notes.append("MAPE could not be computed.")
        failure_reasons.append("MAPE could not be computed (no usable FPE).")
    elif mape_v:
        notes.append(f"MAPE {mape * 100:.2f}% is at or above the 18% limit (non-compliant).")
        failure_reasons.append(
            f"MAPE {mape * 100:.2f}% is at or above the 18% limit (must stay below 18%)."
        )
    if perc95 is None:
        notes.append("PERC95 could not be computed.")
        failure_reasons.append("PERC95 could not be computed.")
    elif p95_v:
        notes.append(f"PERC95 {perc95 * 100:.2f}% is at or above the 30% limit (non-compliant).")
        failure_reasons.append(
            f"PERC95 {perc95 * 100:.2f}% is at or above the 30% limit (must stay below 30%)."
        )
    if day_ok:
        notes.append("Both MAPE and PERC95 are within policy for this day.")
        analysis_summary = (
            f"Compliant: MAPE {mape * 100:.2f}% and PERC95 {perc95 * 100:.2f}% are below policy "
            f"(under {MAPE_MAX_EXCLUSIVE * 100:.0f}% MAPE, under {PERC95_MAX_EXCLUSIVE * 100:.0f}% PERC95)."
            if mape is not None and perc95 is not None
            else "Both MAPE and PERC95 are within policy for this day."
        )
    else:
        analysis_summary = (
            " · ".join(failure_reasons)
            if failure_reasons
            else "Non-compliant: MAPE and/or PERC95 did not meet policy."
        )
    return {
        "mape_max_policy_pct": MAPE_MAX_EXCLUSIVE * 100,
        "perc95_max_policy_pct": PERC95_MAX_EXCLUSIVE * 100,
        "mape_ok": mape_ok,
        "perc95_ok": perc95_ok,
        "day_compliant": day_ok,
        "mape_violation": mape_v,
        "perc95_violation": p95_v,
        "notes": notes,
        "failure_reasons": failure_reasons,
        "analysis_summary": analysis_summary,
    }


def compute_fpe_and_metrics(
    mq_mw: list[float],
    rtd_mw: list[float],
) -> dict[str, Any]:
    if len(mq_mw) != N_INTERVALS or len(rtd_mw) != N_INTERVALS:
        raise ValueError("Internal length mismatch")

    f_load = [0.0] * N_INTERVALS
    for i in range(1, N_INTERVALS):
        f_load[i] = rtd_mw[i - 1]

    g_proj = [(f_load[i] + rtd_mw[i]) / 24.0 for i in range(N_INTERVALS)]
    h_max = max(mq_mw) if mq_mw else 0.0

    fpe: list[float | None] = []
    for i in range(N_INTERVALS):
        if h_max <= 0:
            fpe.append(None)
        else:
            fpe.append(abs((g_proj[i] - mq_mw[i]) / h_max))

    # MAPE: SUM(I)/COUNT(C) — template counts all RTD cells (288)
    n_c = N_INTERVALS
    valid_fpe = [x for x in fpe if x is not None]
    mape = sum(valid_fpe) / n_c if n_c else None

    # PERC95: sorted FPE ascending, h = 0.95*(N+1), Excel TRUNC + fractional
    sorted_f = sorted(x for x in fpe if x is not None)
    n = len(sorted_f)
    perc95 = None
    if n > 0:
        h = 0.95 * (n + 1)
        k = int(h)
        d = h - k
        if k <= 0:
            perc95 = sorted_f[0]
        elif k >= n:
            perc95 = sorted_f[-1]
        else:
            yk = sorted_f[k - 1]
            yk1 = sorted_f[k]
            perc95 = yk + d * (yk1 - yk)

    return {
        "mape": mape,
        "perc95": perc95,
        "n_intervals": N_INTERVALS,
        "n_fpe_sorted": n,
        "max_mq_mw": h_max,
        "fpe": fpe,
        "mq_mw": mq_mw,
        "rtd_mw": rtd_mw,
    }


def _cell_mw(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, str):
        s = val.strip()
        if s in ("-", "—", "N/A", "n/a"):
            return 0.0
        s = s.replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _interval_time_from_cell(val: Any) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                continue
    return None


def _find_rtd_dispatch_sheet(wb: Any) -> Any:
    """Prefer sheet name ``RTD -Actual -Day Ahead`` (ARECO workbook)."""
    names = wb.sheetnames
    for n in names:
        if n.strip().lower() == "rtd -actual -day ahead":
            return wb[n]
    for n in names:
        low = n.lower()
        if "rtd" in low and "actual" in low:
            return wb[n]
    if names:
        return wb[names[0]]
    raise ValueError("Workbook has no sheets.")


def parse_rtd_dispatch_rows_for_day(
    content: bytes,
    trade_date: date,
) -> tuple[list[tuple[datetime, float, float, float]], int]:
    """
    Read ARECO ``RTD -Actual -Day Ahead`` sheet: INTERVAL (time), RTD, ACTUAL, Day Ahead (MW).
    Returns (rows, n_data_rows) where each row is (interval_end_dt, rtd_mw, actual_mw, day_ahead_mw).
    Day-ahead MW is used as the E = MQ series for FPE when MIRF is not used.
    """
    bio = io.BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        sh = _find_rtd_dispatch_sheet(wb)
        out: list[tuple[datetime, float, float, float]] = []
        n_seen = 0
        max_r = sh.max_row or 400
        for r in range(1, max_r + 1):
            t_iv = _interval_time_from_cell(sh.cell(r, 3).value)
            if t_iv is None:
                continue
            n_seen += 1
            rtd = _cell_mw(sh.cell(r, 4).value)
            act = _cell_mw(sh.cell(r, 5).value)
            da = _cell_mw(sh.cell(r, 6).value)
            dt_end = datetime.combine(trade_date, t_iv)
            out.append((dt_end, rtd, act, da))
        if not out:
            raise ValueError(
                "No interval times in column C (expected 5‑minute times, e.g. 06:25:00)."
            )
        return out, n_seen
    finally:
        wb.close()


def fill_mq_from_day_ahead_rows(
    rows: list[tuple[datetime, float, float, float]],
    day: date,
) -> tuple[list[float], int]:
    """Map Day Ahead MW into 288×5‑min slots (full day)."""
    mq = [0.0] * N_INTERVALS
    n = 0
    for dt, _, _, da in rows:
        if dt.date() != day:
            continue
        idx = _interval_index_from_end(dt.time())
        if idx is None:
            continue
        mq[idx] = da
        n += 1
    return mq, n


def analyze_rtd_dispatch_workbook(
    content: bytes,
    filename: str,
    trade_date: date,
    mq_xlsx_bytes: bytes | None = None,
) -> dict[str, Any]:
    """
    RTD ``RTD … Day Ahead`` workbook: RTD and Actual from columns D/E.

    MQ (E in FPE): if ``mq_xlsx_bytes`` is set, DEL MW comes from the MIRF Daily MQ workbook
    (same B13:Y36 logic as MPI+MIRF). Otherwise Day Ahead (column F) is used as MQ — numbers
    may differ from MIRF DEL.
    """
    rows, _n_rows = parse_rtd_dispatch_rows_for_day(content, trade_date)
    if not rows:
        raise ValueError("No rows parsed from workbook.")

    day = trade_date
    compliance = [(dt, rtd, act) for dt, rtd, act, _da in rows]
    if mq_xlsx_bytes is not None:
        mq_mw, sheet_used = load_mq_del_mw_from_xlsx(mq_xlsx_bytes)
        n_mq = N_INTERVALS
        mq_sheet_label = sheet_used
    else:
        mq_mw, n_mq = fill_mq_from_day_ahead_rows(rows, day)
        mq_sheet_label = "Day Ahead (MW) from workbook"

    rtd, actual, n_paste = fill_rtd_actual(compliance, day)
    metrics = compute_fpe_and_metrics(mq_mw, rtd)
    fpe_list = metrics["fpe"]
    metrics["compliance_day"] = day.isoformat()
    metrics["compliance_rows_in_window"] = n_paste
    metrics["mq_sheet"] = mq_sheet_label
    policy = evaluate_nomination_policy(metrics.get("mape"), metrics.get("perc95"))
    analytics = compute_summary_style_analytics(rtd, actual, mq_mw, fpe_list)

    summary = {k: v for k, v in metrics.items() if k not in ("fpe", "mq_mw", "rtd_mw")}
    if summary.get("mape") is not None:
        summary["mape_pct"] = float(summary["mape"]) * 100.0
    if summary.get("perc95") is not None:
        summary["perc95_pct"] = float(summary["perc95"]) * 100.0

    return {
        "summary": summary,
        "policy": policy,
        "analytics": analytics,
        "source": "rtd_dispatch_workbook",
        "filename": filename,
        "mq_intervals_filled": n_mq,
    }


def analyze_uploads(mq_xlsx_bytes: bytes, compliance_csv_bytes: bytes) -> dict[str, Any]:
    mq_mw, sheet_used = load_mq_del_mw_from_xlsx(mq_xlsx_bytes)
    compliance = parse_compliance_csv(compliance_csv_bytes)
    day = _dominant_date(compliance)
    if not day:
        raise ValueError("No compliance rows with a parseable date.")

    rtd, actual, n_paste = fill_rtd_actual(compliance, day)
    metrics = compute_fpe_and_metrics(mq_mw, rtd)
    fpe_list = metrics["fpe"]
    metrics["compliance_day"] = day.isoformat()
    metrics["compliance_rows_in_window"] = n_paste
    metrics["mq_sheet"] = sheet_used
    policy = evaluate_nomination_policy(metrics.get("mape"), metrics.get("perc95"))
    analytics = compute_summary_style_analytics(rtd, actual, mq_mw, fpe_list)

    summary = {k: v for k, v in metrics.items() if k not in ("fpe", "mq_mw", "rtd_mw")}
    if summary.get("mape") is not None:
        summary["mape_pct"] = float(summary["mape"]) * 100.0
    if summary.get("perc95") is not None:
        summary["perc95_pct"] = float(summary["perc95"]) * 100.0

    return {
        "summary": summary,
        "policy": policy,
        "analytics": analytics,
    }
